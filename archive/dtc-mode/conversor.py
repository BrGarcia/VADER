"""
Script Conversor de Arquivos DMP (TRIMM)
Baseado na macro VBA original 'PitchTrimSwitchMonitor'

Funcionalidades:
1. Lê todos os arquivos TRIMM*.DMP da pasta onde o script está executando.
2. Consolida os dados em um único DataFrame Pandas.
3. Adiciona a coluna 'Origem_Arquivo' para rastreabilidade.
4. Gera a coluna 'Hora' convertendo o 'UTC' (que está em milissegundos).
5. Calcula a flag 'Aileron_Test' e 'Elevator_Test' reproduzindo as fórmulas VBA.
6. Salva resultados com nomes padronizados e em ordem de leitura.
"""

import sys
import pandas as pd
import numpy as np
import re
from pathlib import Path

def gerar_prefixo(nome_pasta: str) -> str:
    """Limpa e formata o nome da pasta para ser usado como prefixo padrão."""
    limpo = nome_pasta.upper().replace(" - ", "_").replace(" ", "_").replace("-", "_")
    limpo = re.sub(r'_+', '_', limpo)
    return limpo

def localizar_arquivos_dmp(diretorio: Path) -> list:
    """Localiza todos os arquivos que sigam o padrão TRIMM*.DMP."""
    arquivos = list(diretorio.glob("TRIMM*.DMP"))
    return sorted(arquivos)

def ler_arquivo_dmp(caminho_arquivo: Path) -> pd.DataFrame:
    """Lê um arquivo DMP específico e retorna um DataFrame com as colunas corretas."""
    # Definição das colunas com base na análise da macro VBA
    colunas_originais = [
        "UTC", "Emer_ON", "Emer_SW", "Stick_FWD", "Stick_AFT", 
        "CAS", "TAS", "GS", "BARO", "RALT", "PITCH_ANG", 
        "ROLL_ANG", "AIL_T_POS", "ELEV_T_POS", "RUD_T_POS", 
        "PITCH_MIS", "ROLL_MIS", "Yellow_Zone", "Descarte_Delimitador"
    ]
    
    try:
        df = pd.read_csv(
            caminho_arquivo, 
            sep=";", 
            header=None, 
            names=colunas_originais, 
            index_col=False,
            # Se houver linhas com problemas, ignoramos ou forçamos
            on_bad_lines='skip'
        )
        
        # O último ; gera uma coluna vazia, então a descartamos
        if "Descarte_Delimitador" in df.columns:
            df = df.drop(columns=["Descarte_Delimitador"])
            
        # Adiciona o nome do arquivo de origem na primeira posição
        df.insert(0, "Origem_Arquivo", caminho_arquivo.name)
        
        return df
    except Exception as e:
        print(f"Erro ao ler o arquivo {caminho_arquivo.name}: {e}")
        return pd.DataFrame()

def processar_e_consolidar(lista_dfs: list) -> pd.DataFrame:
    """Consolida a lista de DataFrames e aplica as lógicas da macro VBA."""
    if not lista_dfs:
        return pd.DataFrame()
        
    df_final = pd.concat(lista_dfs, ignore_index=True)
    
    # Verifica se há dados suficientes
    if df_final.empty:
        return df_final
        
    # 1. Cria a coluna Hora a partir do UTC
    # O UTC no arquivo está em milissegundos do dia.
    # Convertendo para datetime usando a origem do epoch do pandas e extraindo string formatada.
    horas_formatadas = pd.to_datetime(df_final["UTC"], unit='ms').dt.strftime('%H:%M:%S.%f')
    # Remove os últimos 3 caracteres para manter .000 (microssegundos viram milissegundos)
    horas_formatadas = horas_formatadas.str[:-3]
    df_final.insert(2, "Hora", horas_formatadas)  # Insere depois do UTC (que é index 1 agora, pois index 0 é Origem)
    
    # 2. Calcula o Threshold de tempo
    # Macro VBA: =(A7-A6)*2 (usando as duas primeiras linhas de tempo, convertido para ms)
    if len(df_final) >= 2:
        threshold_ms = (df_final["UTC"].iloc[1] - df_final["UTC"].iloc[0]) * 2
    else:
        threshold_ms = 0
        
    # 3. Calcula as lógicas de Aileron e Elevator
    # Em pandas, para comparar com a "próxima" linha, usamos shift(-1).
    # Condição: Tempo_Prox - Tempo_Atual > Threshold
    diff_utc = df_final["UTC"].shift(-1) - df_final["UTC"]
    condicao_tempo = diff_utc > threshold_ms
    
    # Aileron_Test: ABS(N_atual - N_prox) > 1 (Coluna N = AIL_T_POS)
    diff_ail = (df_final["AIL_T_POS"] - df_final["AIL_T_POS"].shift(-1)).abs()
    condicao_aileron = condicao_tempo & (diff_ail > 1)
    df_final["Aileron_Test"] = condicao_aileron.astype(int)
    
    # Elevator_Test: ABS(O_atual - O_prox) > 1 (Coluna O = ELEV_T_POS)
    diff_elev = (df_final["ELEV_T_POS"] - df_final["ELEV_T_POS"].shift(-1)).abs()
    condicao_elevator = condicao_tempo & (diff_elev > 1)
    df_final["Elevator_Test"] = condicao_elevator.astype(int)
    
    # Guarda o threshold como atributo para exibir depois
    df_final.attrs["Threshold_ms"] = threshold_ms
    
    return df_final

def formatar_tempo_excel(ms_val):
    """Converte milissegundos num formato de tempo parecido com o do Excel (hh:mm:ss.000)."""
    if pd.isna(ms_val):
        return ""
    td = pd.to_timedelta(ms_val, unit='ms')
    total_seconds = int(td.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    millis = int(td.microseconds / 1000)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d},{millis:03d}"

def aplicar_formatacao_excel(caminho_xlsx: Path, threshold_ms: int, df_consolidado: pd.DataFrame):
    """Aplica formatações visuais e cria o cabeçalho superior replicando 'TrimTimeExtraction'."""
    try:
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    wb = load_workbook(caminho_xlsx)
    ws = wb.active

    # Inserir 5 linhas em branco no topo
    ws.insert_rows(1, 5)

    # Escrever totais na esquerda
    ws['A1'] = 'Threshold'
    ws['B1'] = f"00:00:00,{threshold_ms:03d}"
    ws['A2'] = 'Aileron_Counter'
    ws['B2'] = df_consolidado["Aileron_Test"].sum()
    ws['A3'] = 'Elevator_Counter'
    ws['B3'] = df_consolidado["Elevator_Test"].sum()

    # Mapear as posições exatas dos erros para o bloco da direita
    coluna_atual = 5 # Inicia na Coluna E (índice 5)
    
    # Índices com erro Aileron
    erros_aileron = df_consolidado[df_consolidado['Aileron_Test'] == 1].index.tolist()
    # Índices com erro Elevator
    erros_elevator = df_consolidado[df_consolidado['Elevator_Test'] == 1].index.tolist()
    
    # Processar Extrator de Tempos (TrimTimeExtraction)
    # Lógica: A linha no Excel agora será o index + 7 (1 do index em Python + 1 do cabeçalho que desceu + 5 das linhas inseridas)
    # Porém, a macro aponta a linha *antes* da inserção e *antes* do header. 
    # Para ser exato com o dado, reportamos (index do DataFrame) + 6 como "Linha" (que bate com o número da linha original do Excel com o header na linha 6).
    
    def desenhar_bloco(idx, tipo, coluna_inicio):
        linha_excel = idx + 6
        
        # Pega a linha do erro e a linha seguinte
        try:
            row_erro = df_consolidado.iloc[idx]
            row_seguinte = df_consolidado.iloc[idx+1]
        except IndexError:
            return # Erro na última linha, ignora
            
        pos_col = "AIL_T_POS" if tipo == "Aileron" else "ELEV_T_POS"
        
        c_linha = get_column_letter(coluna_inicio)
        c_tempo = get_column_letter(coluna_inicio + 1)
        c_pos = get_column_letter(coluna_inicio + 2)

        # Cabeçalhos do bloco
        ws[f'{c_linha}1'] = 'Linha'
        ws[f'{c_tempo}1'] = 'Tempo'
        ws[f'{c_pos}1'] = 'Posição'
        
        # Dados do momento do erro
        ws[f'{c_linha}2'] = linha_excel
        ws[f'{c_tempo}2'] = formatar_tempo_excel(row_erro['UTC'])
        ws[f'{c_pos}2'] = row_erro[pos_col]
        
        # Dados da linha seguinte
        ws[f'{c_linha}3'] = linha_excel + 1
        ws[f'{c_tempo}3'] = formatar_tempo_excel(row_seguinte['UTC'])
        ws[f'{c_pos}3'] = row_seguinte[pos_col]
        
        # Legenda inferior
        ws[f'{c_linha}4'] = tipo
        
    for idx in erros_elevator:
        desenhar_bloco(idx, "Elevator", coluna_atual)
        coluna_atual += 3
        
    for idx in erros_aileron:
        desenhar_bloco(idx, "Aileron", coluna_atual)
        coluna_atual += 3

    # Congelar painéis na linha de cabeçalhos (que agora está na linha 6)
    ws.freeze_panes = "A7"
    
    # Aplicar AutoFilter na linha de cabeçalhos
    ws.auto_filter.ref = f"A6:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Estilo de destaque (vermelho claro com texto vermelho escuro - padrão Excel)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006')

    # Mapear colunas de dados (a partir da linha 6)
    colunas_por_nome = {}
    for col in range(1, ws.max_column + 1):
        nome_coluna = ws.cell(row=6, column=col).value
        if nome_coluna:
            colunas_por_nome[nome_coluna] = get_column_letter(col)

    # 1. Colunas de Teste -> Realçar valor 1
    for nome in ["Aileron_Test", "Elevator_Test"]:
        if nome in colunas_por_nome:
            letra = colunas_por_nome[nome]
            range_str = f"{letra}7:{letra}{ws.max_row}"
            ws.conditional_formatting.add(
                range_str,
                CellIsRule(operator='equal', formula=['1'], fill=red_fill, font=red_font)
            )

    # 2. Colunas de Status -> Realçar valor "T"
    for nome in ["Emer_ON", "Emer_SW", "Stick_FWD", "Stick_AFT"]:
        if nome in colunas_por_nome:
            letra = colunas_por_nome[nome]
            range_str = f"{letra}7:{letra}{ws.max_row}"
            ws.conditional_formatting.add(
                range_str,
                CellIsRule(operator='equal', formula=['"T"'], fill=red_fill, font=red_font)
            )
            
    wb.save(caminho_xlsx)

def gerar_relatorio_sintese(diretorio: Path, prefixo: str, qtd_arquivos: int, aileron_count: int, elevator_count: int, threshold_ms: int):
    """Gera um arquivo de relatório consolidado replicando o comportamento do reportWB."""
    caminho_relatorio = diretorio / f"{prefixo}_01_Report_TRIMM.xlsx"
    
    colunas = [
        "Pasta", 
        "Arquivos TRIMM", 
        "Flag atuação AIL:\n1 - Atuação não comandada", 
        "Flag atuação ELEV:\n1 - Atuação não comandada", 
        "Flag de atuação", 
        "Flag de Threshold:\n<= 00:00:00,2 - OK", 
        "Check de Dados:\n0 - OK", 
        "Força Aérea", "Base", "Aeronave", "1P", "2P", "Missão", "Data", "Origem sinal"
    ]
    
    flag_atuacao = 1 if (aileron_count > 0 or elevator_count > 0) else 0
    
    linha = {
        colunas[0]: diretorio.name,
        colunas[1]: qtd_arquivos,
        colunas[2]: aileron_count,
        colunas[3]: elevator_count,
        colunas[4]: flag_atuacao,
        colunas[5]: f"{threshold_ms} ms",
        colunas[6]: 0, 
        colunas[7]: "", colunas[8]: "", colunas[9]: "", colunas[10]: "", 
        colunas[11]: "", colunas[12]: "", colunas[13]: "", colunas[14]: ""
    }
    
    df_report = pd.DataFrame([linha])
    try:
        df_report.to_excel(caminho_relatorio, index=False, engine="openpyxl")
        print(f"-> Relatório de síntese gerado com sucesso em: {caminho_relatorio.name}")
        
        # Ajuste de layout similar ao VBA
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        wb = load_workbook(caminho_relatorio)
        ws = wb.active
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        
        for col in ws.columns:
            letra = col[0].column_letter
            ws.column_dimensions[letra].width = 25
            col[0].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            
        wb.save(caminho_relatorio)
    except Exception as e:
        print(f"Erro ao gerar relatório de síntese: {e}")

def processar_vadr(diretorio: Path, prefixo: str):
    """Busca e formata arquivos VADR (Mishap Time History Data Set)."""
    arquivos_vadr = list(diretorio.glob("*Mishap Time History Data Set.csv"))
    if not arquivos_vadr:
        return
        
    print(f"\nIniciando processamento de VADR. Encontrado(s) {len(arquivos_vadr)} arquivo(s).")
    
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter, column_index_from_string
        import csv
    except ImportError:
        print("Aviso: 'openpyxl' não instalado. Processamento VADR ignorado.")
        return

    for arquivo in arquivos_vadr:
        print(f"- Processando {arquivo.name}...")
        # Usa o prefixo padronizado gerado
        caminho_xlsx = diretorio / f"{prefixo}_02_VADR_Formatado.xlsx"
        
        try:
            wb = Workbook()
            ws = wb.active
            
            # Lê o CSV com o módulo nativo para lidar com números variáveis de colunas por linha
            with open(arquivo, 'r', encoding='utf-8', errors='replace') as f:
                leitor_csv = csv.reader(f, delimiter=',')
                for linha in leitor_csv:
                    ws.append(linha)
            
            # Congelar painéis e AutoFilter
            ws.freeze_panes = "B11"
            # O VBA aplica filtro na linha 10
            ws.auto_filter.ref = f"A10:{get_column_letter(ws.max_column)}{ws.max_row}"
            
            # Larguras específicas
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 17
            
            # Colunas para ocultar baseadas na macro:
            colunas_ocultas = [
                'D', 'E', 'F', 'J', 'K', 'L', 'N', 'O', 'Q', 'S', 'T', 'U', 'W', 'Y',
                'AA', 'AC', 'AE', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL'
            ]
            
            def esconder_intervalo(inicio, fim):
                idx_inicio = column_index_from_string(inicio)
                idx_fim = column_index_from_string(fim)
                for i in range(idx_inicio, idx_fim + 1):
                    col_letra = get_column_letter(i)
                    ws.column_dimensions[col_letra].hidden = True
            
            # Esconde as colunas individuais
            for col in colunas_ocultas:
                ws.column_dimensions[col].hidden = True
                
            # Esconde os intervalos
            esconder_intervalo('AM', 'CC')
            esconder_intervalo('CE', 'CG')
            esconder_intervalo('CK', 'GD')
            esconder_intervalo('GH', 'IX')
            
            print(f"- Salvando arquivo XLSX formatado (isso pode demorar alguns segundos)...")
            wb.save(caminho_xlsx)
            print(f"-> Arquivo VADR formatado salvo como: {caminho_xlsx.name}")
            
        except Exception as e:
            print(f"Erro ao processar VADR {arquivo.name}: {e}")

def processar_subpasta(diretorio_atual: Path):
    print(f"\n{'='*60}\nEntrando na pasta: {diretorio_atual.name}\n{'='*60}")
    
    prefixo_padrao = gerar_prefixo(diretorio_atual.name)
    arquivos_dmp = localizar_arquivos_dmp(diretorio_atual)
    
    flag_disparo = False
    processou_dmp = False
    
    if arquivos_dmp:
        print(f"Foram encontrados {len(arquivos_dmp)} arquivo(s) DMP para processamento.\n")
        
        lista_dfs = []
        for arquivo in arquivos_dmp:
            df = ler_arquivo_dmp(arquivo)
            if not df.empty:
                linhas = len(df)
                print(f"- Lidos {linhas} registros do arquivo {arquivo.name}")
                lista_dfs.append(df)
                
        if lista_dfs:
            processou_dmp = True
            df_consolidado = processar_e_consolidar(lista_dfs)
            total_linhas = len(df_consolidado)
            
            aileron_count = df_consolidado["Aileron_Test"].sum()
            elevator_count = df_consolidado["Elevator_Test"].sum()
            threshold = df_consolidado.attrs.get("Threshold_ms", 0)
            
            # Seta a flag baseada na análise
            if aileron_count > 0 or elevator_count > 0:
                flag_disparo = True
            
            print(f"\nResumo da Consolidação:")
            print(f"-> Total de linhas consolidadas: {total_linhas}")
            print(f"-> Threshold calculado: {threshold} ms")
            print(f"-> Atuações não comandadas (Aileron_Test): {aileron_count}")
            print(f"-> Atuações não comandadas (Elevator_Test): {elevator_count}")
            if flag_disparo:
                print(f"-> ALERTA: Suspeita de disparo detectada!")
            
            print("\nSalvando resultados...")
            caminho_csv = diretorio_atual / f"{prefixo_padrao}_04_TRIMM_Dados_Brutos.csv"
            caminho_xlsx = diretorio_atual / f"{prefixo_padrao}_03_TRIMM_Consolidado.xlsx"
            
            try:
                df_consolidado.to_csv(caminho_csv, index=False, sep=";")
                print(f"-> Salvo com sucesso em: {caminho_csv.name}")
            except Exception as e:
                print(f"Erro ao salvar CSV: {e}")
                
            try:
                df_consolidado.to_excel(caminho_xlsx, index=False, engine="openpyxl")
                print(f"-> Salvo com sucesso em: {caminho_xlsx.name}")
                aplicar_formatacao_excel(caminho_xlsx, threshold, df_consolidado)
                print(f"-> Formatação visual aplicada no Excel com sucesso!")
                gerar_relatorio_sintese(diretorio_atual, prefixo_padrao, len(arquivos_dmp), aileron_count, elevator_count, threshold)
            except Exception as e:
                print(f"Erro ao salvar XLSX: {e}")
        else:
            print("Nenhum dado válido pôde ser extraído dos arquivos DMP.")
    else:
        print("Nenhum arquivo TRIMM*.DMP encontrado nesta pasta.")

    # Executa a formatação dos arquivos VADR, se existirem
    processar_vadr(diretorio_atual, prefixo_padrao)
    
    # Renomeia a pasta baseada na flag
    if processou_dmp:
        sufixo = "_SUSPEITA_DISPARO" if flag_disparo else "_SEM_DISPARO"
        novo_nome = f"{diretorio_atual.name}{sufixo}"
        novo_caminho = diretorio_atual.parent / novo_nome
        
        try:
            diretorio_atual.rename(novo_caminho)
            print(f"\n[!] Pasta rebatizada com sucesso para: {novo_nome}")
        except PermissionError:
            print(f"\n[X] Erro: Não foi possível renomear a pasta para '{novo_nome}'. Verifique se há algum arquivo aberto nela.")
        except Exception as e:
            print(f"\n[X] Erro ao renomear pasta: {e}")
    else:
        print(f"\n[-] Pasta não rebatizada pois nenhum arquivo DMP foi processado.")

def main():
    diretorio_base = Path.cwd()
    print(f"Iniciando Scanner de Voo no diretório base:\n{diretorio_base}\n")
    
    # Localiza todas as subpastas (ignorando pastas ocultas, pacotes offline e que já foram processadas)
    subpastas = [p for p in diretorio_base.iterdir() if p.is_dir() and not p.name.startswith('.') and p.name != 'pacotes_offline']
    
    if not subpastas:
        print("Nenhuma subpasta encontrada para processamento.")
        return
        
    pastas_processadas = 0
    for subpasta in subpastas:
        # Pula as pastas que já foram batizadas em execuções anteriores
        if "_SEM_DISPARO" in subpasta.name or "_SUSPEITA_DISPARO" in subpasta.name:
            print(f"Pulando pasta já analisada: {subpasta.name}")
            continue
            
        processar_subpasta(subpasta)
        pastas_processadas += 1

    print(f"\n{'='*60}")
    print(f"Processo em lote concluído! {pastas_processadas} pasta(s) nova(s) analisada(s).")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()